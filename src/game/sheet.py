import numpy as np
from openpyxl.cell import Cell
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.alignment import Alignment
from openpyxl.workbook import Workbook
import os
import pandas as pd
import uuid

from game.namedata import overlay

outputdir = os.path.dirname(os.getcwd()) + "/output/"

tags = [
  "",
  "",
]

labels = [
]

def makeheader(wsws, mapname, n, cover=False):
  sepfill = PatternFill(patternType='solid', fgColor=Color(rgb='000000'))
  wsws["A1"] = "TAGS"; wsws["A1"].alignment = Alignment(horizontal="center")
  wsws["I1"] = "LEFT"; wsws["M1"] = "RIGHT"; wsws["K1"] = "DATA"
  wsws["A3"] = mapname
  wsws["A4"] = f"{n} GAMES" if n > 1 else f"{n} GAME"

  wsws["K3"] = "NAME"; wsws["K4"] = "AGENT"; wsws["K5"] = "ROLE"; wsws["K6"] = "SUBROLE"; wsws["K7"] = "K / D"
  wsws.column_dimensions["A"].width = 20; wsws.column_dimensions["K"].width = 27
  wsws.column_dimensions["B"].width = 8; wsws.column_dimensions["D"].width = 5
  wsws.column_dimensions["H"].width = 3; wsws.column_dimensions["N"].width = 3
  wsws.column_dimensions["J"].width = 5; wsws.column_dimensions["L"].width = 5
  wsws.column_dimensions["S"].width = 25

  colkeys = ["A", "C", "D", "E", "F", "G", "I", "K", "M", "O", "P", "Q", "R", "S", "T"]
  colov = {
    "LEFT_1": "C", "LEFT_2": "D", "LEFT_3": "E", "LEFT_4": "F", "LEFT_5": "G", 
    "RIGHT_1": "O", "RIGHT_2": "P", "RIGHT_3": "Q", "RIGHT_4": "R", "RIGHT_5": "S", 
    }

  for ov in overlay:  wsws[f"{colov[ov]}1"] = ov
  for col in colkeys:
    if col != "A" and col != "K":  wsws.column_dimensions[col].width = 16
    wsws[f"{col}1"].alignment = Alignment(horizontal="center")
    wsws[f"{col}2"].fill = sepfill
    wsws[f"{col}8"].fill = sepfill
    wsws[f"{col}{9 + len(labels)}"].fill = sepfill

  wsws["A1"].fill = PatternFill(patternType='solid', fgColor=Color(rgb='b99afe'))
  wsws["K1"].fill = PatternFill(patternType='solid', fgColor=Color(rgb='fce600'))
  for rrr in range(1, 111):
    wsws[f"B{rrr}"].fill = sepfill
    wsws[f"H{rrr}"].fill = sepfill
    wsws[f"J{rrr}"].fill = sepfill
    wsws[f"L{rrr}"].fill = sepfill
    wsws[f"N{rrr}"].fill = sepfill
    wsws[f"T{rrr}"].fill = sepfill
    wsws[f"A{rrr}"].alignment = Alignment(horizontal="center")
    wsws[f"K{rrr}"].alignment = Alignment(horizontal="center")
    wsws.row_dimensions[rrr].height = 30

  if cover:
    wsws["A9"] = "1-5: TOP to BOTTOM"

  if not cover:
    for rrr in range(9, 9 + len(labels)):  wsws[f"K{rrr}"] = labels[rrr - 9]

def create_template(maps=[], n=1):
  if type(maps) == str:  maps = [maps]
  u = (uuid.uuid4()).hex
  name = f"{outputdir}valytix_{u}.xlsx"
  wb = Workbook()

  for i in maps:
    ws0 = wb.create_sheet(f"{i}-{str(n)}-COVER-SHEET")
    ws1 = wb.create_sheet(f"{i}-{str(n)}-OVERALL")

    makeheader(ws0, i, n, cover=True)
    for wsws in wb.worksheets[2:]:  makeheader(wsws, i, n)

  wb.save(name)
